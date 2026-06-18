import os
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.vk_friends.formatters import FRIEND_FIELDS, FRIEND_HEADERS_RU, format_cell_value

EXPORT_BATCH_SIZE = 1000
EXPORT_DIR = os.path.abspath(os.path.join(os.getcwd(), ".temp", "vk-friends"))

def _create_sheet(wb: openpyxl.Workbook) -> Any:
    ws = wb.active
    ws.title = "Друзья"
    return ws

def _style_header_row(ws: Any) -> None:
    bold_font = Font(bold=True)
    gray_fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = gray_fill

def _auto_adjust_columns(ws: Any) -> None:
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(14, max_len + 2)

def write_xlsx_file(job_id: str, rows: list[dict[str, Any]]) -> str:
    os.makedirs(EXPORT_DIR, exist_ok=True)
    file_name = f"vk_friends_{job_id}.xlsx"
    file_path = os.path.join(EXPORT_DIR, file_name)

    wb = openpyxl.Workbook()
    ws = _create_sheet(wb)

    headers = [FRIEND_HEADERS_RU.get(field, field) for field in FRIEND_FIELDS]
    ws.append(headers)
    _style_header_row(ws)

    for r in rows:
        row_vals = [format_cell_value(r.get(field)) for field in FRIEND_FIELDS]
        ws.append(row_vals)

    _auto_adjust_columns(ws)

    wb.save(file_path)

    if not os.path.isfile(file_path) or os.path.getsize(file_path) == 0:
        raise RuntimeError(f"Failed to verify file creation: {file_path}. File missing or empty.")

    return file_path
