import os
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.services.ok_friends.formatters import format_cell_value, to_russian_header

EXPORT_BATCH_SIZE = 1000

def get_export_dir() -> str:
    return os.path.abspath(os.path.join(os.getcwd(), settings.ok_friends_export_dir))

def _collect_ordered_keys(rows: list[dict[str, Any]]) -> list[str]:
    seen_keys = set()
    ordered_keys = []
    for row in rows:
        for key in row:
            if key not in seen_keys:
                seen_keys.add(key)
                ordered_keys.append(key)
    return ordered_keys

def _create_sheet(wb: openpyxl.Workbook) -> Any:
    ws = wb.active
    ws.title = "Друзья"
    return ws

def _style_header_row(ws: Any) -> None:
    bold_font = Font(bold=True)
    gray_fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = gray_fill

def _auto_adjust_columns(ws: Any) -> None:
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value or "")
            if len(val) > max_len:
                max_len = len(val)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(14, max_len + 2), 100)

def write_xlsx_file(job_id: str, rows: list[dict[str, Any]]) -> str:
    export_dir = get_export_dir()
    os.makedirs(export_dir, exist_ok=True)
    file_name = f"ok_friends_{job_id}.xlsx"
    file_path = os.path.join(export_dir, file_name)

    if not rows:
        raise ValueError("No data to export")

    wb = openpyxl.Workbook()
    ws = _create_sheet(wb)

    ordered_keys = _collect_ordered_keys(rows)
    headers = [to_russian_header(key) for key in ordered_keys]
    ws.append(headers)
    _style_header_row(ws)

    for r in rows:
        row_vals = [format_cell_value(r.get(key)) for key in ordered_keys]
        ws.append(row_vals)

    _auto_adjust_columns(ws)

    wb.save(file_path)

    if not os.path.isfile(file_path) or os.path.getsize(file_path) == 0:
        raise RuntimeError(f"Failed to verify file creation: {file_path}. File missing or empty.")

    return file_path
