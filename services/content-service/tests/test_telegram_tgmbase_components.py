import io
import sys
from datetime import UTC, datetime
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _service_path import use_service_path

use_service_path()

from app.modules.telegram_tgmbase.exporter import TelegramDlMatchExporter  # noqa: E402
from app.modules.telegram_tgmbase.mapper import TelegramTgmbaseMapper  # noqa: E402
from app.modules.telegram_tgmbase.search import (  # noqa: E402
    TelegramTgmbaseSearchService,
    normalize_tgmbase_query,
)


def test_mapper_builds_search_profile_and_candidate():
    mapper = TelegramTgmbaseMapper()
    user = SimpleNamespace(
        id=1,
        user_id=123456,
        username="CaseUser",
        phone="+79991234567",
        first_name=" Case ",
        last_name=" User ",
        bot=False,
        scam=False,
        premium=True,
        upd_date=None,
    )

    profile = mapper.map_profile(user)
    candidate = mapper.map_candidate(user)

    assert profile["fullName"] == "Case User"
    assert profile["telegramId"] == "123456"
    assert profile["premium"] is True
    assert candidate == {
        "telegramId": "123456",
        "username": "CaseUser",
        "phoneNumber": "+79991234567",
        "fullName": "Case User",
    }


def test_mapper_builds_dl_match_result_with_typed_user_schema():
    mapper = TelegramTgmbaseMapper()
    created_at = datetime(2026, 5, 31, 12, 0, tzinfo=UTC)
    result = SimpleNamespace(
        id=11,
        run_id=7,
        dl_contact_id=22,
        tgmbase_user_id=123,
        strict_telegram_id_match=True,
        username_match=False,
        phone_match=True,
        chat_activity_match=True,
        created_at=created_at,
        dl_contact_snapshot={
            "importFileId": "3",
            "originalFileName": "contacts.csv",
            "telegramId": "123",
            "username": "dl_user",
            "phone": "+79990000000",
            "firstName": "DL",
            "lastName": "User",
            "fullName": "DL User",
            "region": "RU",
            "sourceRowIndex": 5,
        },
        tgmbase_user_snapshot={
            "user_id": "123",
            "username": "base_user",
            "phone": "+79990000000",
            "first_name": "Base",
            "last_name": "User",
            "premium": True,
            "scam": False,
            "bot": False,
            "upd_date": "2026-05-30T10:00:00+00:00",
        },
        chats=[
            SimpleNamespace(chat_type="group", peer_id="42", title="Cases", is_excluded=False),
            SimpleNamespace(chat_type="channel", peer_id="99", title="Excluded", is_excluded=True),
        ],
    )

    mapped = mapper.map_result(result)

    assert mapped["id"] == "11"
    assert mapped["runId"] == "7"
    assert mapped["dlContact"]["username"] == "dl_user"
    assert mapped["dlContact"]["phone"] == "+79990000000"
    assert mapped["user"]["user_id"] == "123"
    assert mapped["user"]["username"] == "base_user"
    assert mapped["user"]["phone"] == "+79990000000"
    assert mapped["user"]["relatedChats"] == [
        {"type": "group", "peer_id": "42", "title": "Cases"}
    ]
    assert mapped["createdAt"] == created_at.isoformat()


def test_search_component_keeps_query_and_summary_contracts():
    search = TelegramTgmbaseSearchService(session=SimpleNamespace(), mapper=TelegramTgmbaseMapper())

    assert normalize_tgmbase_query("@Mixed_User") == {
        "rawValue": "@Mixed_User",
        "normalizedValue": "mixed_user",
        "queryType": "username",
    }
    assert set(search._build_phone_variants("+7 999 123-45-67")) == {
        "+7 999 123-45-67",
        "+79991234567",
        "79991234567",
        "89991234567",
    }
    assert search._build_search_summary(
        [
            {"status": "found"},
            {"status": "not_found"},
            {"status": "ambiguous"},
            {"status": "invalid"},
            {"status": "error"},
        ]
    ) == {
        "total": 5,
        "found": 1,
        "notFound": 1,
        "ambiguous": 1,
        "invalid": 1,
        "error": 1,
    }


@pytest.mark.asyncio
async def test_exporter_writes_match_and_message_sheets():
    exporter = TelegramDlMatchExporter()
    content = await exporter.export_run(
        "7",
        [
            {
                "id": "11",
                "dlContactId": "22",
                "strictTelegramIdMatch": True,
                "usernameMatch": False,
                "phoneMatch": True,
                "dlContact": {
                    "telegramId": "123",
                    "username": "dl_user",
                    "phone": "+79990000000",
                },
                "tgmbaseUserId": "123",
                "user": {
                    "user_id": "123",
                    "username": "base_user",
                    "phone": "+79990000000",
                },
                "chats": [{"type": "group", "peer_id": "42", "title": "Cases"}],
            }
        ],
        {
            "11": [
                {
                    "peerId": "42",
                    "title": "Cases",
                    "messages": [
                        {
                            "messageId": "1",
                            "messageDate": "2026-05-31",
                            "text": "hello",
                        }
                    ],
                }
            ]
        },
    )

    workbook = load_workbook(io.BytesIO(content), data_only=True)

    assert workbook.sheetnames == ["matches", "messages"]
    assert workbook["matches"]["A2"].value == "7"
    assert workbook["matches"]["D2"].value == "telegram_id, phone"
    assert workbook["matches"]["I2"].value == "base_user"
    assert workbook["matches"]["J2"].value == "+79990000000"
    assert workbook["messages"]["D2"].value == "1"
    assert workbook["messages"]["E2"].value == "2026-05-31"
    assert workbook["messages"]["F2"].value == "hello"
