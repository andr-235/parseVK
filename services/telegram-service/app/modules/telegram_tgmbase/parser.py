import io
from datetime import datetime

from openpyxl import load_workbook


class TelegramDlImportRow:
    def __init__(self, **kwargs):
        self.source_row_index = kwargs.get("source_row_index")
        self.telegram_id = kwargs.get("telegram_id")
        self.username = kwargs.get("username")
        self.phone = kwargs.get("phone")
        self.first_name = kwargs.get("first_name")
        self.last_name = kwargs.get("last_name")
        self.description = kwargs.get("description")
        self.region = kwargs.get("region")
        self.date = kwargs.get("date")
        self.channels = kwargs.get("channels")
        self.full_name = kwargs.get("full_name")
        self.address = kwargs.get("address")
        self.vk_url = kwargs.get("vk_url")
        self.email = kwargs.get("email")
        self.telegram_contact = kwargs.get("telegram_contact")
        self.instagram = kwargs.get("instagram")
        self.viber = kwargs.get("viber")
        self.odnoklassniki = kwargs.get("odnoklassniki")
        self.birth_date = kwargs.get("birth_date")
        self.username_extra = kwargs.get("username_extra")
        self.geo = kwargs.get("geo")


class TelegramDlImportParseResult:
    def __init__(self, original_file_name: str, sheet_name: str, contacts: list[TelegramDlImportRow]):
        self.original_file_name = original_file_name
        self.replacement_key = original_file_name
        self.sheet_name = sheet_name
        self.contacts = contacts


class TelegramDlImportParser:
    REQUIRED_HEADERS = ["Id", "Телефон", "Дата"]

    def parse(self, file_content: bytes, original_file_name: str) -> TelegramDlImportParseResult:
        wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
        if not wb.sheetnames:
            raise ValueError("Excel file does not contain any worksheets")
        
        sheet = wb.worksheets[0]
        
        # Читаем первый ряд (заголовки)
        first_row = [cell.value for cell in sheet[1]]
        headers = [self._normalize_header(val) for val in first_row]
        
        # Строим индекс колонок
        header_index = {}
        for idx, h in enumerate(headers):
            if h:
                header_index.setdefault(h, []).append(idx)
                
        # Проверяем наличие обязательных заголовков
        missing = [h for h in self.REQUIRED_HEADERS if h not in header_index]
        if missing:
            raise ValueError(f"Отсутствуют обязательные колонки: {', '.join(missing)}")
            
        contacts = []
        # Читаем со второго ряда
        for row_idx in range(2, sheet.max_row + 1):
            row_cells = [cell.value for cell in sheet[row_idx]]
            if self._is_empty_row(row_cells):
                continue
                
            contact = self._map_row(row_cells, headers, header_index, row_idx)
            contacts.append(contact)
            
        cleaned_file_name = original_file_name.strip()
        
        return TelegramDlImportParseResult(
            original_file_name=cleaned_file_name,
            sheet_name=sheet.title,
            contacts=contacts
        )

    def _is_empty_row(self, row_cells: list) -> bool:
        return all(val is None or str(val).strip() == "" for val in row_cells)

    def _normalize_header(self, val) -> str:
        if val is None:
            return ""
        return str(val).strip()

    def _normalize_cell(self, val) -> str | None:
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.isoformat()
        s = str(val).strip()
        return s if s else None

    def _get_cell(self, row_cells: list, col_idx: int | None) -> str | None:
        if col_idx is None or col_idx >= len(row_cells):
            return None
        return self._normalize_cell(row_cells[col_idx])

    def _get_header_pos(self, header_index: dict, name: str) -> int | None:
        positions = header_index.get(name, [])
        return positions[0] if positions else None

    def _get_required_header_pos(self, header_index: dict, name: str) -> int:
        pos = self._get_header_pos(header_index, name)
        if pos is None:
            raise ValueError(f"Отсутствуют обязательные колонки: {name}")
        return pos

    def _map_row(self, row_cells: list, headers: list, header_index: dict, source_row_index: int) -> TelegramDlImportRow:
        username_positions = header_index.get("Username", [])
        first_username_pos = username_positions[0] if len(username_positions) > 0 else None
        second_username_pos = username_positions[1] if len(username_positions) > 1 else None

        return TelegramDlImportRow(
            source_row_index=source_row_index,
            telegram_id=self._get_cell(row_cells, self._get_required_header_pos(header_index, "Id")) or "",
            username=self._get_cell(row_cells, first_username_pos),
            phone=self._get_cell(row_cells, self._get_required_header_pos(header_index, "Телефон")),
            first_name=self._get_cell(row_cells, self._get_header_pos(header_index, "Имя")),
            last_name=self._get_cell(row_cells, self._get_header_pos(header_index, "Фамилия")),
            description=self._get_cell(row_cells, self._get_header_pos(header_index, "Описание")),
            region=self._get_cell(row_cells, self._get_header_pos(header_index, "Регион")),
            date=self._get_cell(row_cells, self._get_required_header_pos(header_index, "Дата")),
            channels=self._get_cell(row_cells, self._get_header_pos(header_index, "Каналы")),
            full_name=self._get_cell(row_cells, self._get_header_pos(header_index, "ФИО")),
            address=self._get_cell(row_cells, self._get_header_pos(header_index, "Адрес")),
            vk_url=self._get_cell(row_cells, self._get_header_pos(header_index, "Вконтакте")),
            email=self._get_cell(row_cells, self._get_header_pos(header_index, "Почта")),
            telegram_contact=self._get_cell(row_cells, self._get_header_pos(header_index, "Телеграм")),
            instagram=self._get_cell(row_cells, self._get_header_pos(header_index, "Инстаграм")),
            viber=self._get_cell(row_cells, self._get_header_pos(header_index, "Viber")),
            odnoklassniki=self._get_cell(row_cells, self._get_header_pos(header_index, "Одноклассники")),
            birth_date=self._get_cell(row_cells, self._get_header_pos(header_index, "Дата рождения")),
            username_extra=self._get_cell(row_cells, second_username_pos),
            geo=self._get_cell(row_cells, self._get_header_pos(header_index, "Гео")),
        )
